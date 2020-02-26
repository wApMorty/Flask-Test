import cv2
import numpy as np
from openpyxl import load_workbook
import os

# Données

NbProtoA = "INCONNU"
NbProtoB = "INCONNU"

NomExcel = 'ImageAnalise.xlsx'

CWD = os.getcwd()
PATH = "S:\\EME2020\\Flask-Test\\appli\\static\\images\\"

workbook = load_workbook(os.path.join(CWD, "analyseFlocs", "Template_vierge.xlsx"))
worksheet = workbook.active

i=0

# Récuperation des images dans le dossier

for file in os.listdir(PATH):
    if ((file.endswith('.jpg')) or (file.endswith('.png'))):
        
        img1 = cv2.imread(PATH + file)

        # Split RGB

        b,g,r = cv2.split(img1)

        # Binarisation avec l'image rouge et modifiant les 3 dernière ligne en vide

        ret,img2 = cv2.threshold(r,180,255,cv2.THRESH_BINARY)
        img2[img2.shape[0]-1] = 255
        img2[img2.shape[0]-2] = 255
        img2[img2.shape[0]-3] = 255

        # 3 dilatations

        kernel = np.ones((3, 3), np.uint8)
        # Séparé en 3 pour comparaison
        img3 = cv2.erode(img2, kernel, iterations = 1)
        img4 = cv2.erode(img3, kernel, iterations = 1)
        img5 = cv2.erode(img4, kernel, iterations = 1)

        # Calcul du rapport

        pixelsFlocs = np.count_nonzero(img5 == 0)
        pourcentage = (int)((pixelsFlocs/(img1.size/3))*100)

        # Remplissage Excel

        worksheet.cell(row=8+i, column=1).value = file
        worksheet.cell(row=8+i, column=3).value = NbProtoA
        worksheet.cell(row=8+i, column=4).value = NbProtoB
        worksheet.cell(row=8+i, column=6).value = pourcentage

        i=i+1

workbook.save(PATH + NomExcel)
workbook.close()